"""Small helper to build simple styled .xlsx reports (vendor balances, meal
counts, etc.) without repeating openpyxl boilerplate in every view. Kept
separate from the quotation exporter, which fills the planner's own template
instead of building a sheet from scratch."""
import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def build_simple_workbook(title, headers, rows, totals_row=None, money_columns=None):
    """money_columns: 1-indexed column numbers to render with thousands
    separators (e.g. [5] for a 5th "Monto" column)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31] or "Reporte"

    ws.append([title])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    if len(headers) > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.append([])

    ws.append(headers)
    header_row = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="212529")
        cell.alignment = Alignment(horizontal="center")

    money_columns = money_columns or []
    for row in rows:
        ws.append(row)
        for col in money_columns:
            ws.cell(row=ws.max_row, column=col).number_format = "#,##0.00"

    if totals_row:
        ws.append(totals_row)
        for col in range(1, len(totals_row) + 1):
            ws.cell(row=ws.max_row, column=col).font = Font(bold=True)
        for col in money_columns:
            ws.cell(row=ws.max_row, column=col).number_format = "#,##0.00"

    for col_idx, col_cells in enumerate(ws.columns, start=1):
        values = [str(c.value) for c in col_cells if c.value is not None]
        length = max((len(v) for v in values), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(length + 2, 10), 45)

    return wb


def workbook_response(wb, filename):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
