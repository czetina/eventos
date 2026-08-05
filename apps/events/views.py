import json
from datetime import date as dt_date, time as dt_time

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db.models import Max, Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone, translation
from django.utils.translation import gettext_lazy as _

from apps.notes.forms import NoteForm

from . import importers
from .forms import (
    EventAdvanceForm, EventForm, EventSectionTypeForm, EventSessionForm, EventTeamMemberForm,
    ExpenseForm, GuestImportForm, GuestQuickEntryForm, InvoiceForm, InvoiceItemForm, MealCountForm,
    ProcessionalEntryForm, QuotationForm, QuotationItemForm, SeatingTableForm, SessionImportForm,
    TableGuestForm, TableGuestMoveForm, WeddingPartyMemberForm, WeddingPartyListMergeForm,
    WeddingPartyListTypeForm, WeddingTableTypeForm,
)
from .models import (
    Event, EventAdvance, EventSectionType, EventSession, EventTeamMember, Expense, Invoice,
    InvoiceItem, MealCount, ProcessionalEntry, Quotation, QuotationItem, SeatingTable, TableGuest,
    WeddingPartyMember, WeddingPartyListType, WeddingTableType, create_default_section_types,
    create_default_table_types, create_default_wedding_party_list_types,
)


def events_visible_to(user):
    """Scope events by tenant + role: admins/planners see the whole company,
    supervisors/encargados only see events where they are on the team."""
    if not user.company:
        return Event.objects.none()
    qs = Event.objects.filter(company=user.company)
    if user.can_manage_events:
        return qs
    return qs.filter(team_members__user=user).distinct()


def get_event_or_403(user, pk):
    event = get_object_or_404(Event, pk=pk)
    if event.company_id != user.company_id:
        raise PermissionDenied
    if not user.can_manage_events and not event.team_members.filter(user=user).exists():
        raise PermissionDenied(_("No perteneces al equipo de este evento."))
    return event


@login_required
def event_list(request):
    events = events_visible_to(request.user).order_by("-event_date")
    return render(request, "events/event_list.html", {"events": events})


@login_required
def event_create(request):
    if not request.user.can_manage_events:
        raise PermissionDenied(_("Solo planificadores o administradores pueden crear eventos."))
    if request.method == "POST":
        form = EventForm(request.POST, company=request.user.company)
        if form.is_valid():
            event = form.save(commit=False)
            event.company = request.user.company
            event.created_by = request.user
            event.save()
            if request.user.role:
                EventTeamMember.objects.get_or_create(
                    event=event, user=request.user, defaults={"role": request.user.role},
                )
            messages.success(request, _(
                "Evento creado correctamente. Ahora puedes importar el guion/tareas desde Excel."
            ))
            return redirect("tasks:import", event_pk=event.pk)
    else:
        form = EventForm(company=request.user.company)
    return render(request, "events/event_form.html", {"form": form, "is_new": True})


@login_required
def event_edit(request, pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    if request.method == "POST":
        form = EventForm(request.POST, instance=event, company=request.user.company)
        if form.is_valid():
            form.save()
            messages.success(request, _("Evento actualizado."))
            return redirect("events:detail", pk=event.pk)
    else:
        form = EventForm(instance=event, company=request.user.company)
    return render(request, "events/event_form.html", {"form": form, "is_new": False, "event": event})


@login_required
def event_delete(request, pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied(_("Solo planificadores o administradores pueden eliminar eventos."))
    if request.method == "POST":
        name = event.name
        event.delete()
        messages.success(request, _("Evento '%(name)s' eliminado, junto con sus tareas, itinerario y notas.") % {
            "name": name
        })
        return redirect("events:list")
    return render(request, "events/event_confirm_delete.html", {"event": event})


@login_required
def event_detail(request, pk):
    event = get_event_or_403(request.user, pk)
    tasks = event.tasks.select_related("assigned_to", "supervisor")
    if not request.user.can_manage_events:
        tasks = tasks.filter(assigned_to=request.user)
    return render(request, "events/event_detail.html", {
        "event": event,
        "sessions": event.sessions.all(),
        "team_members": event.team_members.select_related("user", "reports_to__user"),
        "tasks": tasks[:10],
        "task_total": event.tasks.count(),
        "vendors": event.event_vendors.select_related("vendor"),
        "notes": event.notes.all()[:5],
        "files": event.files.all()[:5],
        "note_form": NoteForm(),
    })


@login_required
def event_session_create(request, pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    create_default_section_types(event.company)
    if request.method == "POST":
        form = EventSessionForm(request.POST, company=event.company)
        if form.is_valid():
            session = form.save(commit=False)
            session.event = event
            session.save()
            messages.success(request, _("Actividad agregada al itinerario."))
            return redirect("events:detail", pk=event.pk)
    else:
        form = EventSessionForm(initial={"date": event.event_date}, company=event.company)
    return render(request, "events/session_form.html", {"form": form, "event": event, "is_new": True})


@login_required
def event_session_edit(request, pk, session_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    session = get_object_or_404(EventSession, pk=session_pk, event=event)
    if request.method == "POST":
        form = EventSessionForm(request.POST, instance=session, company=event.company)
        if form.is_valid():
            form.save()
            messages.success(request, _("Actividad del itinerario actualizada."))
            return redirect("events:detail", pk=event.pk)
    else:
        form = EventSessionForm(instance=session, company=event.company)
    return render(request, "events/session_form.html", {"form": form, "event": event, "is_new": False})


@login_required
def section_type_list(request):
    if not request.user.can_manage_events:
        raise PermissionDenied(_("No tienes permiso para gestionar las secciones del itinerario."))
    create_default_section_types(request.user.company)
    section_types = EventSectionType.objects.filter(company=request.user.company)
    return render(request, "events/section_type_list.html", {"section_types": section_types})


@login_required
def section_type_create(request):
    if not request.user.can_manage_events:
        raise PermissionDenied(_("No tienes permiso para gestionar las secciones del itinerario."))
    if request.method == "POST":
        form = EventSectionTypeForm(request.POST)
        if form.is_valid():
            section_type = form.save(commit=False)
            section_type.company = request.user.company
            section_type.save()
            messages.success(request, _("Sección creada correctamente."))
            return redirect("events:section_type_list")
    else:
        form = EventSectionTypeForm()
    return render(request, "events/section_type_form.html", {"form": form, "is_new": True})


@login_required
def section_type_edit(request, pk):
    if not request.user.can_manage_events:
        raise PermissionDenied(_("No tienes permiso para gestionar las secciones del itinerario."))
    section_type = get_object_or_404(EventSectionType, pk=pk, company=request.user.company)
    if request.method == "POST":
        form = EventSectionTypeForm(request.POST, instance=section_type)
        if form.is_valid():
            form.save()
            messages.success(request, _("Sección actualizada correctamente."))
            return redirect("events:section_type_list")
    else:
        form = EventSectionTypeForm(instance=section_type)
    return render(request, "events/section_type_form.html", {
        "form": form, "is_new": False, "section_type": section_type,
    })


@login_required
def section_type_delete(request, pk):
    if not request.user.can_manage_events:
        raise PermissionDenied(_("No tienes permiso para gestionar las secciones del itinerario."))
    section_type = get_object_or_404(EventSectionType, pk=pk, company=request.user.company)
    if request.method == "POST":
        if section_type.sessions.exists():
            messages.error(request, _("No se puede eliminar una sección que todavía tiene actividades asignadas."))
        else:
            section_type.delete()
            messages.success(request, _("Sección eliminada."))
    return redirect("events:section_type_list")


@login_required
def event_session_delete(request, pk, session_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    session = get_object_or_404(EventSession, pk=session_pk, event=event)
    if request.method == "POST":
        session.delete()
        messages.success(request, _("Actividad eliminada del itinerario."))
    return redirect("events:detail", pk=event.pk)


@login_required
def event_session_bulk_delete(request, pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    if request.method == "POST":
        session_ids = request.POST.getlist("session_ids")
        deleted, _details = EventSession.objects.filter(event=event, pk__in=session_ids).delete()
        if deleted:
            messages.success(request, _("%(count)s actividades del itinerario eliminadas.") % {"count": deleted})
        else:
            messages.error(request, _("No seleccionaste ninguna actividad."))
    return redirect("events:detail", pk=event.pk)


def _row_to_session_payload(row):
    return {
        "title": row.title,
        "notes": row.notes,
        "venue_name": row.venue_name,
        "start_time": row.start_time.strftime("%H:%M") if row.start_time else "",
        "time_is_carried_over": row.time_is_carried_over,
    }


@login_required
def event_session_import(request, pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied(_("Solo planificadores o administradores pueden importar el itinerario."))

    if request.method == "POST":
        form = SessionImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                rows = importers.parse_minuto_a_minuto(form.cleaned_data["file"])
            except ValueError as exc:
                messages.error(request, str(exc))
                return render(request, "events/session_import.html", {"form": form, "event": event})

            if not rows:
                messages.error(request, _("No se encontraron filas para importar en ese archivo."))
                return render(request, "events/session_import.html", {"form": form, "event": event})

            payload = [_row_to_session_payload(r) for r in rows]

            out_of_order = False
            last_time = None
            for row in payload:
                if row["start_time"]:
                    if last_time and row["start_time"] < last_time:
                        out_of_order = True
                    last_time = row["start_time"]

            return render(request, "events/session_import_preview.html", {
                "event": event,
                "rows": payload,
                "payload_json": json.dumps(payload),
                "out_of_order": out_of_order,
            })
    else:
        form = SessionImportForm()
    return render(request, "events/session_import.html", {"form": form, "event": event})


@login_required
def event_session_import_confirm(request, pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied(_("Solo planificadores o administradores pueden importar el itinerario."))
    if request.method != "POST":
        return redirect("events:session_import", pk=event.pk)

    try:
        rows = json.loads(request.POST.get("payload", "[]"))
    except json.JSONDecodeError:
        messages.error(request, _("No se pudo leer la información a importar."))
        return redirect("events:session_import", pk=event.pk)

    next_order = (event.sessions.aggregate(Max("order"))["order__max"] or 0) + 1
    fallback_time = event.start_time or dt_time(0, 0)
    created = 0
    for row in rows:
        start_time = dt_time.fromisoformat(row["start_time"]) if row.get("start_time") else fallback_time
        EventSession.objects.create(
            event=event,
            title=row["title"],
            venue_name=row.get("venue_name", ""),
            date=event.event_date,
            start_time=start_time,
            notes=row.get("notes", ""),
            order=next_order,
        )
        next_order += 1
        created += 1

    messages.success(request, _("Se importaron %(count)s actividades al itinerario.") % {"count": created})
    return redirect("events:detail", pk=event.pk)


@login_required
def event_team(request, pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    if request.method == "POST":
        form = EventTeamMemberForm(request.POST, company=request.user.company, event=event)
        if form.is_valid():
            member = form.save(commit=False)
            member.event = event
            member.save()
            messages.success(request, _("Miembro agregado al equipo del evento."))
            return redirect("events:team", pk=event.pk)
    else:
        form = EventTeamMemberForm(company=request.user.company, event=event)
    return render(request, "events/event_team.html", {
        "event": event,
        "form": form,
        "team_members": event.team_members.select_related("user", "reports_to__user"),
    })


@login_required
def event_team_remove(request, pk, member_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    member = get_object_or_404(EventTeamMember, pk=member_pk, event=event)
    if request.method == "POST":
        member.delete()
        messages.success(request, _("Miembro removido del equipo."))
    return redirect("events:team", pk=event.pk)


@login_required
def event_wedding_party(request, pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    create_default_wedding_party_list_types(event.company)
    if request.method == "POST":
        form = WeddingPartyMemberForm(request.POST, company=event.company)
        if form.is_valid():
            member = form.save(commit=False)
            member.event = event
            member.save()
            messages.success(request, _("Persona agregada a la lista."))
            return redirect("events:wedding_party", pk=event.pk)
    else:
        form = WeddingPartyMemberForm(company=event.company)
    members = event.wedding_party_members.select_related("list_type")
    list_types = list(WeddingPartyListType.objects.filter(company=event.company))
    for list_type in list_types:
        list_type.member_list = members.filter(list_type=list_type)
    return render(request, "events/wedding_party.html", {
        "event": event,
        "form": form,
        "list_types": list_types,
    })


@login_required
def event_wedding_party_edit(request, pk, member_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    member = get_object_or_404(WeddingPartyMember, pk=member_pk, event=event)
    if request.method == "POST":
        form = WeddingPartyMemberForm(request.POST, instance=member, company=event.company)
        if form.is_valid():
            form.save()
            messages.success(request, _("Registro actualizado."))
            return redirect("events:wedding_party", pk=event.pk)
    else:
        form = WeddingPartyMemberForm(instance=member, company=event.company)
    return render(request, "events/wedding_party_form.html", {"form": form, "event": event, "member": member})


@login_required
def event_wedding_party_remove(request, pk, member_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    member = get_object_or_404(WeddingPartyMember, pk=member_pk, event=event)
    if request.method == "POST":
        member.delete()
        messages.success(request, _("Registro eliminado."))
    return redirect("events:wedding_party", pk=event.pk)


@login_required
def event_wedding_party_list_clear(request, pk, list_type_pk):
    """Removes every person *this event* has in a list — the list category
    itself (WeddingPartyListType) is shared company-wide, so it's left in
    place for other events that may still be using it; only delete it via
    'Mantenimiento de listas' once no event has anyone left in it."""
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    list_type = get_object_or_404(WeddingPartyListType, pk=list_type_pk, company=event.company)
    if request.method == "POST":
        deleted, _details = WeddingPartyMember.objects.filter(event=event, list_type=list_type).delete()
        messages.success(request, _("Se eliminaron %(count)s personas de la lista '%(name)s' en este evento.") % {
            "count": deleted, "name": list_type.name,
        })
    return redirect("events:wedding_party", pk=event.pk)


@login_required
def wedding_party_search_guests(request, pk):
    """Alternative to typing a name by hand: search the event's seating chart
    (plan de mesas) and add a matching guest to one or several cortejo lists
    at once, copying their name and table — one WeddingPartyMember row per
    list checked, since a person can be in more than one list (e.g. Damas
    and Discursos) but each list only holds one FK per row."""
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    create_default_wedding_party_list_types(event.company)
    list_types = list(WeddingPartyListType.objects.filter(company=event.company))
    query = request.GET.get("q", "").strip()
    guests = []
    if query:
        guests = list(
            event.all_guests.select_related("table").filter(
                Q(first_name__icontains=query) | Q(last_name__icontains=query) | Q(family__icontains=query)
            ).order_by("first_name", "last_name")
        )

    if request.method == "POST":
        created = 0
        posted_guests = event.all_guests.filter(pk__in=request.POST.getlist("guest_id")).select_related("table")
        for guest in posted_guests:
            for list_type in list_types:
                if request.POST.get(f"list_{guest.pk}_{list_type.pk}") != "on":
                    continue
                next_order = (
                    WeddingPartyMember.objects.filter(event=event, list_type=list_type)
                    .aggregate(Max("order"))["order__max"] or 0
                ) + 1
                WeddingPartyMember.objects.create(
                    event=event, list_type=list_type, name=guest.name, quantity=1,
                    order=next_order, table_number=guest.table.table_number if guest.table else "",
                )
                created += 1
        if created:
            messages.success(request, _("Se agregaron %(count)s registros al cortejo.") % {"count": created})
        else:
            messages.info(request, _("No se seleccionó ninguna lista para las personas marcadas."))
        return redirect(f"{reverse('events:wedding_party_search_guests', args=[event.pk])}?q={query}")

    return render(request, "events/wedding_party_search_guests.html", {
        "event": event, "list_types": list_types, "guests": guests, "query": query,
    })


@login_required
def wedding_party_type_list(request):
    if not request.user.can_manage_events:
        raise PermissionDenied(_("No tienes permiso para gestionar las listas del cortejo."))
    create_default_wedding_party_list_types(request.user.company)
    list_types = WeddingPartyListType.objects.filter(company=request.user.company)
    return render(request, "events/wedding_party_type_list.html", {"list_types": list_types})


@login_required
def wedding_party_type_merge(request, pk):
    """Combines two or more existing lists into a brand-new one, scoped to
    this event only — every person *this event* has in the selected source
    lists is moved (not duplicated) into the new list. List categories are
    shared company-wide, but the people in them are per-event, so people
    other events happen to have in those same lists are left untouched."""
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied(_("No tienes permiso para gestionar las listas del cortejo."))
    if request.method == "POST":
        form = WeddingPartyListMergeForm(request.POST, company=event.company, event=event)
        if form.is_valid():
            source_lists = form.cleaned_data["source_lists"]
            next_order = (
                WeddingPartyListType.objects.filter(company=event.company)
                .aggregate(Max("order"))["order__max"] or 0
            ) + 1
            new_list = WeddingPartyListType.objects.create(
                company=event.company, name=form.cleaned_data["name"], order=next_order,
            )
            moved = WeddingPartyMember.objects.filter(
                event=event, list_type__in=source_lists
            ).update(list_type=new_list)
            messages.success(request, _("Se unieron %(count)s personas en la nueva lista '%(name)s'.") % {
                "count": moved, "name": new_list.name,
            })
            return redirect("events:wedding_party", pk=event.pk)
    else:
        form = WeddingPartyListMergeForm(company=event.company, event=event)
    return render(request, "events/wedding_party_type_merge.html", {"form": form, "event": event})


@login_required
def wedding_party_type_create(request):
    if not request.user.can_manage_events:
        raise PermissionDenied(_("No tienes permiso para gestionar las listas del cortejo."))
    if request.method == "POST":
        form = WeddingPartyListTypeForm(request.POST)
        if form.is_valid():
            list_type = form.save(commit=False)
            list_type.company = request.user.company
            list_type.save()
            messages.success(request, _("Lista creada correctamente."))
            return redirect("events:wedding_party_type_list")
    else:
        form = WeddingPartyListTypeForm()
    return render(request, "events/wedding_party_type_form.html", {"form": form, "is_new": True})


@login_required
def wedding_party_type_edit(request, pk):
    if not request.user.can_manage_events:
        raise PermissionDenied(_("No tienes permiso para gestionar las listas del cortejo."))
    list_type = get_object_or_404(WeddingPartyListType, pk=pk, company=request.user.company)
    if request.method == "POST":
        form = WeddingPartyListTypeForm(request.POST, instance=list_type)
        if form.is_valid():
            form.save()
            messages.success(request, _("Lista actualizada correctamente."))
            return redirect("events:wedding_party_type_list")
    else:
        form = WeddingPartyListTypeForm(instance=list_type)
    return render(request, "events/wedding_party_type_form.html", {
        "form": form, "is_new": False, "list_type": list_type,
    })


@login_required
def wedding_party_type_delete(request, pk):
    if not request.user.can_manage_events:
        raise PermissionDenied(_("No tienes permiso para gestionar las listas del cortejo."))
    list_type = get_object_or_404(WeddingPartyListType, pk=pk, company=request.user.company)
    if request.method == "POST":
        if list_type.members.exists():
            messages.error(request, _("No se puede eliminar una lista que todavía tiene personas asignadas."))
        else:
            list_type.delete()
            messages.success(request, _("Lista eliminada."))
    return redirect("events:wedding_party_type_list")


@login_required
def table_type_list(request):
    if not request.user.can_manage_events:
        raise PermissionDenied(_("No tienes permiso para gestionar los tipos de mesa."))
    create_default_table_types(request.user.company)
    table_types = WeddingTableType.objects.filter(company=request.user.company)
    return render(request, "events/table_type_list.html", {"table_types": table_types})


@login_required
def table_type_create(request):
    if not request.user.can_manage_events:
        raise PermissionDenied(_("No tienes permiso para gestionar los tipos de mesa."))
    if request.method == "POST":
        form = WeddingTableTypeForm(request.POST)
        if form.is_valid():
            table_type = form.save(commit=False)
            table_type.company = request.user.company
            table_type.save()
            messages.success(request, _("Tipo de mesa creado correctamente."))
            return redirect("events:table_type_list")
    else:
        form = WeddingTableTypeForm()
    return render(request, "events/table_type_form.html", {"form": form, "is_new": True})


@login_required
def table_type_edit(request, pk):
    if not request.user.can_manage_events:
        raise PermissionDenied(_("No tienes permiso para gestionar los tipos de mesa."))
    table_type = get_object_or_404(WeddingTableType, pk=pk, company=request.user.company)
    if request.method == "POST":
        form = WeddingTableTypeForm(request.POST, instance=table_type)
        if form.is_valid():
            form.save()
            messages.success(request, _("Tipo de mesa actualizado correctamente."))
            return redirect("events:table_type_list")
    else:
        form = WeddingTableTypeForm(instance=table_type)
    return render(request, "events/table_type_form.html", {
        "form": form, "is_new": False, "table_type": table_type,
    })


@login_required
def table_type_delete(request, pk):
    if not request.user.can_manage_events:
        raise PermissionDenied(_("No tienes permiso para gestionar los tipos de mesa."))
    table_type = get_object_or_404(WeddingTableType, pk=pk, company=request.user.company)
    if request.method == "POST":
        if table_type.tables.exists():
            messages.error(request, _("No se puede eliminar un tipo de mesa que todavía está en uso."))
        else:
            table_type.delete()
            messages.success(request, _("Tipo de mesa eliminado."))
    return redirect("events:table_type_list")


@login_required
def event_seating_chart(request, pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    create_default_table_types(event.company)
    if request.method == "POST":
        form = SeatingTableForm(request.POST, company=event.company)
        if form.is_valid():
            table = form.save(commit=False)
            table.event = event
            table.save()
            messages.success(request, _("Mesa agregada."))
            return redirect("events:seating_chart", pk=event.pk)
    else:
        form = SeatingTableForm(company=event.company)
    tables = event.seating_tables.select_related("table_type").prefetch_related("guests")
    guest_form = TableGuestForm()
    move_form = TableGuestMoveForm(event=event)
    unassigned_guests = event.all_guests.filter(table__isnull=True).order_by("first_name", "last_name")
    return render(request, "events/seating_chart.html", {
        "event": event, "form": form, "tables": tables,
        "guest_form": guest_form, "move_form": move_form,
        "unassigned_guests": unassigned_guests,
    })


@login_required
def event_seating_table_edit(request, pk, table_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    table = get_object_or_404(SeatingTable, pk=table_pk, event=event)
    if request.method == "POST":
        form = SeatingTableForm(request.POST, instance=table, company=event.company)
        if form.is_valid():
            form.save()
            messages.success(request, _("Mesa actualizada."))
            return redirect("events:seating_chart", pk=event.pk)
    else:
        form = SeatingTableForm(instance=table, company=event.company)
    return render(request, "events/seating_table_form.html", {"form": form, "event": event, "table": table})


@login_required
def event_seating_table_remove(request, pk, table_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    table = get_object_or_404(SeatingTable, pk=table_pk, event=event)
    if request.method == "POST":
        table.delete()
        messages.success(request, _("Mesa eliminada."))
    return redirect("events:seating_chart", pk=event.pk)


def _sync_guest_speech(guest, event):
    """Keeps TableGuest.gives_speech in sync with a WeddingPartyMember in the
    company's 'Discursos' list — created/renamed/removed automatically so the
    seating chart and Cortejo nupcial never disagree about who is speaking."""
    if guest.gives_speech:
        if guest.speech_member_id:
            member = guest.speech_member
            changed = []
            if member.name != guest.name:
                member.name = guest.name
                changed.append("name")
            if member.table_number != guest.table.table_number:
                member.table_number = guest.table.table_number
                changed.append("table_number")
            if changed:
                member.save(update_fields=changed)
        else:
            types = create_default_wedding_party_list_types(event.company)
            discursos = types["Discursos"]
            next_order = (
                WeddingPartyMember.objects.filter(event=event, list_type=discursos)
                .aggregate(Max("order"))["order__max"] or 0
            ) + 1
            member = WeddingPartyMember.objects.create(
                event=event, list_type=discursos, name=guest.name, quantity=1,
                order=next_order, table_number=guest.table.table_number,
            )
            guest.speech_member = member
            guest.save(update_fields=["speech_member"])
    elif guest.speech_member_id:
        guest.speech_member.delete()
        guest.speech_member = None
        guest.save(update_fields=["speech_member"])


@login_required
def event_table_guest_add(request, pk, table_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    table = get_object_or_404(SeatingTable, pk=table_pk, event=event)
    if request.method == "POST":
        form = TableGuestForm(request.POST)
        if form.is_valid():
            next_order = (table.guests.aggregate(Max("order"))["order__max"] or 0) + 1
            guest = form.save(commit=False)
            guest.event = event
            guest.table = table
            guest.order = next_order
            guest.save()
            _sync_guest_speech(guest, event)
            messages.success(request, _("Invitado agregado."))
        else:
            messages.error(request, _("No se pudo agregar el invitado: revisa el nombre."))
    return redirect("events:seating_chart", pk=event.pk)


@login_required
def event_table_guest_edit(request, pk, table_pk, guest_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    table = get_object_or_404(SeatingTable, pk=table_pk, event=event)
    guest = get_object_or_404(TableGuest, pk=guest_pk, table=table)
    if request.method == "POST":
        form = TableGuestForm(request.POST, instance=guest)
        if form.is_valid():
            form.save()
            _sync_guest_speech(guest, event)
            messages.success(request, _("Invitado actualizado."))
            return redirect("events:seating_chart", pk=event.pk)
    else:
        form = TableGuestForm(instance=guest)
    return render(request, "events/table_guest_form.html", {"form": form, "event": event, "table": table, "guest": guest})


@login_required
def event_table_guest_remove(request, pk, table_pk, guest_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    table = get_object_or_404(SeatingTable, pk=table_pk, event=event)
    guest = get_object_or_404(TableGuest, pk=guest_pk, table=table)
    if request.method == "POST":
        if guest.speech_member_id:
            guest.speech_member.delete()
        guest.delete()
        messages.success(request, _("Invitado eliminado."))
    return redirect("events:seating_chart", pk=event.pk)


@login_required
def event_table_guest_toggle_speech(request, pk, table_pk, guest_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    table = get_object_or_404(SeatingTable, pk=table_pk, event=event)
    guest = get_object_or_404(TableGuest, pk=guest_pk, table=table)
    if request.method == "POST":
        guest.gives_speech = not guest.gives_speech
        guest.save(update_fields=["gives_speech"])
        _sync_guest_speech(guest, event)
        if guest.gives_speech:
            messages.success(request, _("Invitado agregado a Cortejo nupcial › Discursos."))
        else:
            messages.success(request, _("Invitado quitado de Cortejo nupcial › Discursos."))
    return redirect("events:seating_chart", pk=event.pk)


@login_required
def event_table_guest_move(request, pk, table_pk, guest_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    table = get_object_or_404(SeatingTable, pk=table_pk, event=event)
    guest = get_object_or_404(TableGuest, pk=guest_pk, table=table)
    if request.method == "POST":
        form = TableGuestMoveForm(request.POST, event=event)
        if form.is_valid():
            guest.table = form.cleaned_data["table"]
            guest.save(update_fields=["table"])
            messages.success(request, _("Invitado movido de mesa."))
    return redirect("events:seating_chart", pk=event.pk)


@login_required
def event_guest_assign_table(request, pk, guest_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    guest = get_object_or_404(TableGuest, pk=guest_pk, event=event)
    if request.method == "POST":
        form = TableGuestMoveForm(request.POST, event=event)
        if form.is_valid():
            table = form.cleaned_data["table"]
            next_order = (table.guests.aggregate(Max("order"))["order__max"] or 0) + 1
            guest.table = table
            guest.order = next_order
            guest.save(update_fields=["table", "order"])
            messages.success(request, _("Invitado asignado a mesa %(table)s.") % {"table": table.table_number})
    return redirect("events:seating_chart", pk=event.pk)


_SEXO_LABELS = {code: str(label) for code, label in TableGuest.SEXO_CHOICES}
_INVITA_LABELS = {code: str(label) for code, label in TableGuest.INVITA_CHOICES}


def _row_to_guest_payload(row):
    return {
        "first_name": row.first_name,
        "last_name": row.last_name,
        "name": row.name,
        "table_number": row.table_number,
        "invita": row.invita,
        "invita_display": _INVITA_LABELS.get(row.invita, ""),
        "sexo": row.sexo,
        "sexo_display": _SEXO_LABELS.get(row.sexo, ""),
        "family": row.family,
        "rsvp": row.rsvp,
        "main_dish": row.main_dish,
        "dietary_restrictions": row.dietary_restrictions,
        "notes": row.notes,
        "relationship": row.relationship,
    }


@login_required
def event_guest_import(request, pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied(_("Solo planificadores o administradores pueden importar invitados."))

    if request.method == "POST":
        form = GuestImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                rows = importers.parse_guest_list(form.cleaned_data["file"])
            except ValueError as exc:
                messages.error(request, str(exc))
                return render(request, "events/guest_import.html", {"form": form, "event": event})

            if not rows:
                messages.error(request, _("No se encontraron filas para importar en ese archivo."))
                return render(request, "events/guest_import.html", {"form": form, "event": event})

            payload = [_row_to_guest_payload(r) for r in rows]
            existing_tables = set(event.seating_tables.values_list("table_number", flat=True))
            new_tables = sorted(
                {r.table_number for r in rows if r.table_number} - existing_tables,
                key=lambda n: (0, int(n)) if n.isdigit() else (1, n),
            )
            unassigned_count = sum(1 for r in rows if not r.table_number)
            existing_count = event.all_guests.count()

            return render(request, "events/guest_import_preview.html", {
                "event": event,
                "rows": payload,
                "payload_json": json.dumps(payload),
                "new_tables": new_tables,
                "unassigned_count": unassigned_count,
                "existing_count": existing_count,
            })
    else:
        form = GuestImportForm()
    return render(request, "events/guest_import.html", {"form": form, "event": event})


@login_required
def event_guest_import_confirm(request, pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied(_("Solo planificadores o administradores pueden importar invitados."))
    if request.method != "POST":
        return redirect("events:guest_import", pk=event.pk)

    try:
        rows = json.loads(request.POST.get("payload", "[]"))
    except json.JSONDecodeError:
        messages.error(request, _("No se pudo leer la información a importar."))
        return redirect("events:guest_import", pk=event.pk)

    replaced = event.all_guests.count()
    event.all_guests.all().delete()

    table_types = create_default_table_types(event.company)
    default_type = table_types["Redonda"]
    tables_by_number = {t.table_number: t for t in event.seating_tables.all()}

    created = 0
    for row in rows:
        table = None
        table_number = row.get("table_number", "").strip()
        if table_number:
            table = tables_by_number.get(table_number)
            if table is None:
                table = SeatingTable.objects.create(
                    event=event, table_number=table_number, table_type=default_type,
                )
                tables_by_number[table_number] = table
        next_order = (table.guests.aggregate(Max("order"))["order__max"] or 0) + 1 if table else 0
        TableGuest.objects.create(
            event=event, table=table,
            first_name=row["first_name"], last_name=row.get("last_name", ""),
            invita=row.get("invita", ""), sexo=row.get("sexo", ""),
            family=row.get("family", ""), rsvp=row.get("rsvp", ""),
            main_dish=row.get("main_dish", ""), dietary_restrictions=row.get("dietary_restrictions", ""),
            notes=row.get("notes", ""), relationship=row.get("relationship", ""),
            order=next_order,
        )
        created += 1

    messages.success(request, _("Se importaron %(count)s invitados correctamente.") % {"count": created})
    if replaced:
        messages.info(request, _("Se borraron %(count)s invitados existentes antes de importar.") % {"count": replaced})
    return redirect("events:seating_chart", pk=event.pk)


@login_required
def event_guest_quick_entry(request, pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    table_types = create_default_table_types(event.company)
    if request.method == "POST":
        form = GuestQuickEntryForm(request.POST)
        if form.is_valid():
            table = None
            table_number = form.cleaned_data["table_number"].strip()
            if table_number:
                table, tbl_created = SeatingTable.objects.get_or_create(
                    event=event, table_number=table_number,
                    defaults={"table_type": table_types["Redonda"]},
                )
            next_order = (table.guests.aggregate(Max("order"))["order__max"] or 0) + 1 if table else 0
            guest = TableGuest.objects.create(
                event=event, table=table,
                first_name=form.cleaned_data["first_name"], last_name=form.cleaned_data["last_name"],
                sexo=form.cleaned_data["sexo"], invita=form.cleaned_data["invita"],
                order=next_order,
            )
            messages.success(request, _("Invitado agregado: %(name)s.") % {"name": guest.name})
            return redirect("events:guest_quick_entry", pk=event.pk)
    else:
        form = GuestQuickEntryForm()
    recent_guests = event.all_guests.select_related("table").order_by("-id")[:15]
    return render(request, "events/guest_quick_entry.html", {
        "event": event, "form": form, "recent_guests": recent_guests,
    })


def _guest_sort_key(guest):
    if guest.table:
        tn = guest.table.table_number
        mesa_key = (0, int(tn)) if tn.isdigit() else (1, tn)
    else:
        mesa_key = (2, "")
    return mesa_key + (guest.first_name.lower(), guest.last_name.lower())


def _seating_report_filters(request, event):
    """Reads mesa/familia/sexo/rol from the querystring, applies them to the
    event's guests, and returns (filtered guests as a list, dict of active
    filter values, dict of the option lists to populate the filter selects)."""
    mesa = request.GET.get("mesa", "")
    familia = request.GET.get("familia", "")
    sexo = request.GET.get("sexo", "")
    rol = request.GET.get("rol", "")

    guests = event.all_guests.select_related("table")
    if mesa == "sin":
        guests = guests.filter(table__isnull=True)
    elif mesa:
        guests = guests.filter(table_id=mesa)
    if familia:
        guests = guests.filter(family=familia)
    if sexo:
        guests = guests.filter(sexo=sexo)
    if rol:
        guests = guests.filter(relationship=rol)

    guests = sorted(guests, key=_guest_sort_key)

    all_guests = event.all_guests.select_related("table")
    family_options = sorted({g.family for g in all_guests if g.family})
    role_options = sorted({g.relationship for g in all_guests if g.relationship})
    table_options = sorted(
        event.seating_tables.all(), key=lambda t: (0, int(t.table_number)) if t.table_number.isdigit() else (1, t.table_number)
    )

    return guests, {"mesa": mesa, "familia": familia, "sexo": sexo, "rol": rol}, {
        "family_options": family_options, "role_options": role_options, "table_options": table_options,
    }


def _seating_report_resumen(guests):
    def count(pred):
        return sum(1 for g in guests if pred(g))

    return {
        "total": len(guests),
        "confirmados": count(lambda g: "confirmad" in g.rsvp.lower()),
        "pendientes": count(lambda g: "pendient" in g.rsvp.lower()),
        "sin_rsvp": count(lambda g: not g.rsvp),
        "hombres": count(lambda g: g.sexo == TableGuest.SEXO_HOMBRE),
        "mujeres": count(lambda g: g.sexo == TableGuest.SEXO_MUJER),
        "ninos": count(lambda g: g.sexo == TableGuest.SEXO_NINO),
        "ninas": count(lambda g: g.sexo == TableGuest.SEXO_NINA),
        "novio": count(lambda g: g.invita == TableGuest.INVITA_NOVIO),
        "novia": count(lambda g: g.invita == TableGuest.INVITA_NOVIA),
        "ambos": count(lambda g: g.invita == TableGuest.INVITA_AMBOS),
        "sin_mesa": count(lambda g: not g.table_id),
    }


@login_required
def report_seating_chart(request, pk):
    event = get_event_or_403(request.user, pk)
    lang = request.GET.get("lang") or getattr(request, "LANGUAGE_CODE", None) or translation.get_language()
    tables = event.seating_tables.select_related("table_type")
    total_capacity = sum(t.capacity for t in tables)

    with translation.override(lang):
        guests, active_filters, filter_options = _seating_report_filters(request, event)
        resumen = _seating_report_resumen(guests)
        return render(request, "events/report_seating_chart.html", {
            "event": event, "lang": lang, "guests": guests,
            "total_capacity": total_capacity, "resumen": resumen,
            "active_filters": active_filters, "sexo_choices": TableGuest.SEXO_CHOICES,
            **filter_options,
        })


@login_required
def report_seating_chart_excel(request, pk):
    from .xlsx_export import build_simple_workbook, workbook_response
    import openpyxl
    from openpyxl.styles import Font

    event = get_event_or_403(request.user, pk)
    lang = request.GET.get("lang") or getattr(request, "LANGUAGE_CODE", None) or translation.get_language()

    with translation.override(lang):
        guests, _active_filters, _filter_options = _seating_report_filters(request, event)
        resumen = _seating_report_resumen(guests)

        rows = [[
            g.table.table_number if g.table else "", g.first_name, g.last_name, g.family,
            g.get_sexo_display() if g.sexo else "", g.get_invita_display() if g.invita else "",
            g.relationship, g.rsvp, g.main_dish, g.dietary_restrictions, g.notes,
        ] for g in guests]

        wb = build_simple_workbook(
            f"{_('Plan de mesas')} - {event.name}",
            [
                str(_("Mesa")), str(_("Nombre")), str(_("Apellido")), str(_("Familia")),
                str(_("Sexo")), str(_("Invita")), str(_("Rol")), str(_("RSVP")),
                str(_("Plato principal")), str(_("Restricciones")), str(_("Anotaciones")),
            ],
            rows,
        )
        ws = wb.active
        ws.append([])
        ws.append([str(_("Resumen"))])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
        summary_lines = [
            (_("Total de invitados"), resumen["total"]),
            (_("Confirmados"), resumen["confirmados"]),
            (_("Pendientes"), resumen["pendientes"]),
            (_("Sin RSVP"), resumen["sin_rsvp"]),
            (_("Hombres"), resumen["hombres"]),
            (_("Mujeres"), resumen["mujeres"]),
            (_("Niños"), resumen["ninos"]),
            (_("Niñas"), resumen["ninas"]),
            (_("Lado del novio"), resumen["novio"]),
            (_("Lado de la novia"), resumen["novia"]),
            (_("Ambos lados"), resumen["ambos"]),
            (_("Sin mesa asignada"), resumen["sin_mesa"]),
        ]
        for label, value in summary_lines:
            ws.append([str(label), value])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    return workbook_response(wb, f"plan_de_mesas_{event.pk}.xlsx")


def _salida_diagram_order(entries_salida):
    """For the salida (exit) diagram, the top of the church is the door and
    the bottom is the altar: order 1 (first to walk out) belongs at the top
    near the door, and walking order increases toward the altar. Order 0 is
    reserved for a fixed reference at the altar (e.g. the officiant, who
    doesn't walk out with the rest), so it's always placed last/closest to
    the altar regardless of the numeric order of everyone else."""
    return sorted(entries_salida, key=lambda entry: (entry.order == 0, entry.order))


@login_required
def event_processional(request, pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    if request.method == "POST":
        form = ProcessionalEntryForm(request.POST)
        if form.is_valid():
            entry = form.save(commit=False)
            entry.event = event
            entry.save()
            messages.success(request, _("Entrada agregada al planograma."))
            return redirect("events:processional", pk=event.pk)
    else:
        next_order = (
            event.processional_entries.filter(phase=ProcessionalEntry.PHASE_ENTRADA)
            .aggregate(Max("order"))["order__max"] or 0
        ) + 1
        form = ProcessionalEntryForm(initial={"order": next_order, "phase": ProcessionalEntry.PHASE_ENTRADA})
    public_url = request.build_absolute_uri(reverse("events:processional_public", args=[event.share_token]))
    entries_entrada = event.processional_entries.filter(phase=ProcessionalEntry.PHASE_ENTRADA)
    entries_salida = event.processional_entries.filter(phase=ProcessionalEntry.PHASE_SALIDA)
    return render(request, "events/processional.html", {
        "event": event, "form": form,
        "entries_entrada": entries_entrada,
        "entries_salida": entries_salida,
        "diagram_entries_entrada": entries_entrada.order_by("order"),
        "diagram_entries_salida": _salida_diagram_order(entries_salida),
        "public_url": public_url,
    })


@login_required
def event_processional_edit(request, pk, entry_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    entry = get_object_or_404(ProcessionalEntry, pk=entry_pk, event=event)
    if request.method == "POST":
        form = ProcessionalEntryForm(request.POST, instance=entry)
        if form.is_valid():
            form.save()
            messages.success(request, _("Entrada del planograma actualizada."))
            return redirect("events:processional", pk=event.pk)
    else:
        form = ProcessionalEntryForm(instance=entry)
    return render(request, "events/processional_form.html", {"form": form, "event": event, "entry": entry})


@login_required
def event_processional_remove(request, pk, entry_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    entry = get_object_or_404(ProcessionalEntry, pk=entry_pk, event=event)
    if request.method == "POST":
        entry.delete()
        messages.success(request, _("Entrada eliminada del planograma."))
    return redirect("events:processional", pk=event.pk)


def processional_public(request, token):
    """Read-only, no-login view of the church processional diagram — the link a
    planner shares with the bride and, from there, over WhatsApp."""
    event = get_object_or_404(Event, share_token=token)
    return render(request, "events/processional_public.html", {
        "event": event,
        "is_public": True,
        "diagram_entries_entrada": event.processional_entries.filter(
            phase=ProcessionalEntry.PHASE_ENTRADA
        ).order_by("order"),
        "diagram_entries_salida": _salida_diagram_order(
            event.processional_entries.filter(phase=ProcessionalEntry.PHASE_SALIDA)
        ),
    })


@login_required
def event_meals(request, pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    if request.method == "POST":
        form = MealCountForm(request.POST)
        if form.is_valid():
            meal = form.save(commit=False)
            meal.event = event
            meal.save()
            messages.success(request, _("Conteo de comida agregado."))
            return redirect("events:meals", pk=event.pk)
    else:
        form = MealCountForm()
    meals = event.meal_counts.all()
    team_meals = meals.filter(group=MealCount.GROUP_TEAM)
    vendor_meals = meals.filter(group=MealCount.GROUP_VENDOR)
    contingency_meals = meals.filter(group=MealCount.GROUP_CONTINGENCY)
    return render(request, "events/meals.html", {
        "event": event,
        "form": form,
        "team_meals": team_meals,
        "vendor_meals": vendor_meals,
        "contingency_meals": contingency_meals,
        "team_total": team_meals.aggregate(Sum("amount"))["amount__sum"] or 0,
        "vendor_total": vendor_meals.aggregate(Sum("amount"))["amount__sum"] or 0,
        "contingency_total": contingency_meals.aggregate(Sum("amount"))["amount__sum"] or 0,
        "grand_total": meals.aggregate(Sum("amount"))["amount__sum"] or 0,
    })


@login_required
def event_meals_edit(request, pk, meal_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    meal = get_object_or_404(MealCount, pk=meal_pk, event=event)
    if request.method == "POST":
        form = MealCountForm(request.POST, instance=meal)
        if form.is_valid():
            form.save()
            messages.success(request, _("Conteo de comida actualizado."))
            return redirect("events:meals", pk=event.pk)
    else:
        form = MealCountForm(instance=meal)
    return render(request, "events/meals_form.html", {"form": form, "event": event, "meal": meal})


@login_required
def event_meals_remove(request, pk, meal_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    meal = get_object_or_404(MealCount, pk=meal_pk, event=event)
    if request.method == "POST":
        meal.delete()
        messages.success(request, _("Conteo de comida eliminado."))
    return redirect("events:meals", pk=event.pk)


@login_required
def event_expenses(request, pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    if request.method == "POST":
        form = ExpenseForm(request.POST, request.FILES)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.event = event
            expense.created_by = request.user
            expense.save()
            messages.success(request, _("Gasto agregado."))
            return redirect("events:expenses", pk=event.pk)
    else:
        form = ExpenseForm()
    return render(request, "events/expenses.html", {
        "event": event, "form": form, "expenses": event.expenses.all(),
        "advance_form": EventAdvanceForm(), "advances": event.advances.all(),
    })


@login_required
def event_expense_edit(request, pk, expense_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    expense = get_object_or_404(Expense, pk=expense_pk, event=event)
    if request.method == "POST":
        old_document_name = expense.document.name if expense.document else None
        form = ExpenseForm(request.POST, request.FILES, instance=expense)
        if form.is_valid():
            form.save()
            new_document_name = expense.document.name if expense.document else None
            if old_document_name and old_document_name != new_document_name:
                expense.document.storage.delete(old_document_name)
            messages.success(request, _("Gasto actualizado."))
            return redirect("events:expenses", pk=event.pk)
    else:
        form = ExpenseForm(instance=expense)
    return render(request, "events/expense_form.html", {"form": form, "event": event, "expense": expense})


@login_required
def event_expense_remove(request, pk, expense_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    expense = get_object_or_404(Expense, pk=expense_pk, event=event)
    if request.method == "POST":
        if expense.document:
            expense.document.delete(save=False)
        expense.delete()
        messages.success(request, _("Gasto eliminado."))
    return redirect("events:expenses", pk=event.pk)


@login_required
def event_advance_add(request, pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    if request.method == "POST":
        form = EventAdvanceForm(request.POST)
        if form.is_valid():
            advance = form.save(commit=False)
            advance.event = event
            advance.created_by = request.user
            advance.save()
            messages.success(request, _("Anticipo agregado."))
    return redirect("events:expenses", pk=event.pk)


@login_required
def event_advance_edit(request, pk, advance_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    advance = get_object_or_404(EventAdvance, pk=advance_pk, event=event)
    if request.method == "POST":
        form = EventAdvanceForm(request.POST, instance=advance)
        if form.is_valid():
            form.save()
            messages.success(request, _("Anticipo actualizado."))
            return redirect("events:expenses", pk=event.pk)
    else:
        form = EventAdvanceForm(instance=advance)
    return render(request, "events/advance_form.html", {"form": form, "event": event, "advance": advance})


@login_required
def event_advance_remove(request, pk, advance_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    advance = get_object_or_404(EventAdvance, pk=advance_pk, event=event)
    if request.method == "POST":
        advance.delete()
        messages.success(request, _("Anticipo eliminado."))
    return redirect("events:expenses", pk=event.pk)


@login_required
def quotation_list(request, pk):
    event = get_event_or_403(request.user, pk)
    return render(request, "events/quotation_list.html", {
        "event": event, "quotations": event.quotations.all(),
    })


@login_required
def quotation_create(request, pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    if request.method == "POST":
        form = QuotationForm(request.POST)
        if form.is_valid():
            quotation = form.save(commit=False)
            quotation.event = event
            quotation.created_by = request.user
            quotation.correlative = event.quotations.count() + 1
            quotation.save()
            messages.success(request, _("Cotización creada."))
            return redirect("events:quotation_detail", pk=event.pk, quotation_pk=quotation.pk)
    else:
        form = QuotationForm(initial={
            "realization_date": event.event_date, "client_name": event.client_name, "activity": event.name,
        })
    return render(request, "events/quotation_form.html", {"form": form, "event": event, "is_new": True})


@login_required
def quotation_edit(request, pk, quotation_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    quotation = get_object_or_404(Quotation, pk=quotation_pk, event=event)
    if request.method == "POST":
        form = QuotationForm(request.POST, instance=quotation)
        if form.is_valid():
            form.save()
            messages.success(request, _("Cotización actualizada."))
            return redirect("events:quotation_detail", pk=event.pk, quotation_pk=quotation.pk)
    else:
        form = QuotationForm(instance=quotation)
    return render(request, "events/quotation_form.html", {
        "form": form, "event": event, "is_new": False, "quotation": quotation,
    })


@login_required
def quotation_delete(request, pk, quotation_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    quotation = get_object_or_404(Quotation, pk=quotation_pk, event=event)
    if request.method == "POST":
        quotation.delete()
        messages.success(request, _("Cotización eliminada."))
        return redirect("events:quotation_list", pk=event.pk)
    return redirect("events:quotation_detail", pk=event.pk, quotation_pk=quotation.pk)


@login_required
def quotation_detail(request, pk, quotation_pk):
    event = get_event_or_403(request.user, pk)
    quotation = get_object_or_404(Quotation, pk=quotation_pk, event=event)
    if request.method == "POST":
        if not request.user.can_manage_events:
            raise PermissionDenied
        next_order = (quotation.items.aggregate(Max("order"))["order__max"] or 0) + 1
        form = QuotationItemForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.quotation = quotation
            item.order = next_order
            item.save()
            messages.success(request, _("Línea agregada a la cotización."))
            return redirect("events:quotation_detail", pk=event.pk, quotation_pk=quotation.pk)
    else:
        form = QuotationItemForm()
    return render(request, "events/quotation_detail.html", {
        "event": event, "quotation": quotation, "items": quotation.items.all(), "form": form,
    })


@login_required
def quotation_item_delete(request, pk, quotation_pk, item_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    quotation = get_object_or_404(Quotation, pk=quotation_pk, event=event)
    item = get_object_or_404(QuotationItem, pk=item_pk, quotation=quotation)
    if request.method == "POST":
        item.delete()
        messages.success(request, _("Línea eliminada."))
    return redirect("events:quotation_detail", pk=event.pk, quotation_pk=quotation.pk)


@login_required
def quotation_item_edit(request, pk, quotation_pk, item_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    quotation = get_object_or_404(Quotation, pk=quotation_pk, event=event)
    item = get_object_or_404(QuotationItem, pk=item_pk, quotation=quotation)
    if request.method == "POST":
        form = QuotationItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, _("Línea actualizada."))
            return redirect("events:quotation_detail", pk=event.pk, quotation_pk=quotation.pk)
    else:
        form = QuotationItemForm(instance=item)
    return render(request, "events/quotation_item_form.html", {
        "form": form, "event": event, "quotation": quotation, "item": item,
    })


@login_required
def quotation_export_excel(request, pk, quotation_pk, client_version=False):
    import openpyxl
    from pathlib import Path

    from openpyxl.styles import Font

    event = get_event_or_403(request.user, pk)
    quotation = get_object_or_404(Quotation, pk=quotation_pk, event=event)
    items = list(quotation.items.all())

    template_path = Path(__file__).resolve().parent / "xlsx_templates" / "modelo_cotizacion.xlsx"
    wb = openpyxl.load_workbook(template_path)

    # The "td" sheet is the planner's own unrelated personal reconciliation
    # tab from the original example file — never part of the quotation.
    if "td" in wb.sheetnames:
        del wb["td"]

    ws = wb["COTIZACION"]

    # Push everything down 2 rows to make room for the correlativo header line.
    ws.insert_rows(1, amount=2)

    # These cells hold the planner's own internal markup/profit scratch-work
    # from the original example file — not part of what was asked for, and
    # their formulas would break once rows are inserted for >16 items, so
    # they're cleared in every export rather than carried over incorrectly.
    for coord in ("H26", "I26", "H28", "I28", "I29", "C28", "J19"):
        ws[coord] = None

    # Fixed leftover text from the example file, cleared in every export.
    ws["C8"] = None  # was the "PARA ENTREGAR..." note
    ws["A23"] = None  # was the fixed "SOLO CAJA" label

    ws["A1"] = f"No. {quotation.correlative}"
    ws["A1"].font = Font(bold=True, color="FFFF0000", size=13)

    first_row = 10
    last_template_row = 25
    total_row_template = 26

    ws["C4"] = quotation.realization_date
    ws["C5"] = quotation.client_name
    ws["C6"] = quotation.activity
    ws["F5"] = float(quotation.exchange_rate)

    for coord in ("B4", "C4", "B5", "C5", "B6", "C6"):
        cell = ws[coord]
        cell.font = Font(
            name=cell.font.name, size=cell.font.size, bold=True,
            italic=cell.font.italic, color=cell.font.color,
        )

    extra_rows = max(0, len(items) - (last_template_row - first_row + 1))
    if extra_rows:
        ws.insert_rows(total_row_template, amount=extra_rows)

    last_item_row = last_template_row + extra_rows
    total_row = total_row_template + extra_rows

    src_e = ws[f"E{last_template_row}"]
    for row in range(first_row, last_item_row + 1):
        ws[f"B{row}"] = None
        ws[f"C{row}"] = None
        ws[f"D{row}"] = None
        ws[f"E{row}"] = None
        e_cell = ws[f"E{row}"]
        e_cell.number_format = src_e.number_format
        ws[f"F{row}"] = f"=E{row}/$F$5"
        ws[f"F{row}"].number_format = src_e.number_format

    for idx, item in enumerate(items):
        row = first_row + idx
        ws[f"B{row}"] = item.vendor_name
        ws[f"C{row}"] = item.detail
        ws[f"D{row}"] = item.quantity
        ws[f"E{row}"] = float(item.value_dop)

    ws["C9"] = f'=C6&" USD "&TEXT(F{total_row},"#,###.##")'
    ws[f"E{total_row}"] = f"=SUM(E{first_row}:E{last_item_row})"
    ws[f"F{total_row}"] = f"=E{total_row}/$F$5"

    if client_version:
        ws["B8"] = None
        for row in range(first_row, last_item_row + 1):
            ws[f"B{row}"] = None

        # These reference the internal exchange-rate mechanics — kept for the
        # formulas that depend on them, just visually hidden from the client.
        for coord in ("E5", "F5", "E6", "F6"):
            cell = ws[coord]
            cell.font = Font(
                name=cell.font.name, size=cell.font.size, bold=cell.font.bold,
                italic=cell.font.italic, color="FFFFFFFF",
            )

        ws[f"C{total_row}"] = None  # no "GRAN TOTAL" label for the client copy

    filename = f"cotizacion_{quotation.pk}_{'cliente' if client_version else 'completa'}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _next_invoice_number(company):
    last_numbers = [
        int(n) for n in Invoice.objects.filter(event__company=company).values_list("invoice_number", flat=True)
        if n.isdigit()
    ]
    return str(max(last_numbers) + 1) if last_numbers else "1"


@login_required
def invoice_list(request, pk):
    event = get_event_or_403(request.user, pk)
    return render(request, "events/invoice_list.html", {
        "event": event, "invoices": event.invoices.all(),
    })


@login_required
def invoice_create(request, pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    if request.method == "POST":
        form = InvoiceForm(request.POST)
        if form.is_valid():
            invoice = form.save(commit=False)
            invoice.event = event
            invoice.created_by = request.user
            invoice.save()
            messages.success(request, _("Factura creada."))
            return redirect("events:invoice_detail", pk=event.pk, invoice_pk=invoice.pk)
    else:
        form = InvoiceForm(initial={
            "invoice_number": _next_invoice_number(event.company),
            "date": timezone.localdate(),
            "bill_to_name": event.client_name,
            "job_name": event.name,
        })
    return render(request, "events/invoice_form.html", {"form": form, "event": event, "is_new": True})


@login_required
def invoice_edit(request, pk, invoice_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    invoice = get_object_or_404(Invoice, pk=invoice_pk, event=event)
    if invoice.is_locked:
        messages.error(request, _("Esta factura está anulada o pagada; no se puede modificar."))
        return redirect("events:invoice_detail", pk=event.pk, invoice_pk=invoice.pk)
    if request.method == "POST":
        form = InvoiceForm(request.POST, instance=invoice)
        if form.is_valid():
            form.save()
            messages.success(request, _("Factura actualizada."))
            return redirect("events:invoice_detail", pk=event.pk, invoice_pk=invoice.pk)
    else:
        form = InvoiceForm(instance=invoice)
    return render(request, "events/invoice_form.html", {
        "form": form, "event": event, "is_new": False, "invoice": invoice,
    })


@login_required
def invoice_delete(request, pk, invoice_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    invoice = get_object_or_404(Invoice, pk=invoice_pk, event=event)
    if request.method == "POST":
        invoice.delete()
        messages.success(request, _("Factura eliminada."))
        return redirect("events:invoice_list", pk=event.pk)
    return redirect("events:invoice_detail", pk=event.pk, invoice_pk=invoice.pk)


@login_required
def invoice_void(request, pk, invoice_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    invoice = get_object_or_404(Invoice, pk=invoice_pk, event=event)
    if request.method == "POST":
        invoice.status = Invoice.STATUS_VOID
        invoice.voided_at = timezone.now()
        invoice.voided_by = request.user
        invoice.save(update_fields=["status", "voided_at", "voided_by"])
        messages.success(request, _("Factura anulada."))
    return redirect("events:invoice_detail", pk=event.pk, invoice_pk=invoice.pk)


@login_required
def invoice_mark_paid(request, pk, invoice_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    invoice = get_object_or_404(Invoice, pk=invoice_pk, event=event)
    if request.method == "POST":
        received_date = request.POST.get("payment_received_date") or timezone.localdate()
        invoice.payment_received = True
        invoice.payment_received_date = received_date
        invoice.save(update_fields=["payment_received", "payment_received_date"])
        messages.success(request, _("Factura marcada como pagada."))
    return redirect("events:invoice_detail", pk=event.pk, invoice_pk=invoice.pk)


@login_required
def invoice_unmark_paid(request, pk, invoice_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    invoice = get_object_or_404(Invoice, pk=invoice_pk, event=event)
    if request.method == "POST":
        invoice.payment_received = False
        invoice.payment_received_date = None
        invoice.save(update_fields=["payment_received", "payment_received_date"])
        messages.success(request, _("Se quitó la marca de pagada."))
    return redirect("events:invoice_detail", pk=event.pk, invoice_pk=invoice.pk)


@login_required
def invoice_detail(request, pk, invoice_pk):
    event = get_event_or_403(request.user, pk)
    invoice = get_object_or_404(Invoice, pk=invoice_pk, event=event)
    if request.method == "POST":
        if not request.user.can_manage_events:
            raise PermissionDenied
        if invoice.is_locked:
            messages.error(request, _("Esta factura está anulada o pagada; no se puede modificar."))
            return redirect("events:invoice_detail", pk=event.pk, invoice_pk=invoice.pk)
        next_order = (invoice.items.aggregate(Max("order"))["order__max"] or 0) + 1
        form = InvoiceItemForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.invoice = invoice
            item.order = next_order
            item.save()
            messages.success(request, _("Línea agregada a la factura."))
            return redirect("events:invoice_detail", pk=event.pk, invoice_pk=invoice.pk)
    else:
        form = InvoiceItemForm()
    return render(request, "events/invoice_detail.html", {
        "event": event, "invoice": invoice, "items": invoice.items.all(), "form": form,
    })


@login_required
def invoice_item_edit(request, pk, invoice_pk, item_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    invoice = get_object_or_404(Invoice, pk=invoice_pk, event=event)
    item = get_object_or_404(InvoiceItem, pk=item_pk, invoice=invoice)
    if invoice.is_locked:
        messages.error(request, _("Esta factura está anulada o pagada; no se puede modificar."))
        return redirect("events:invoice_detail", pk=event.pk, invoice_pk=invoice.pk)
    if request.method == "POST":
        form = InvoiceItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, _("Línea actualizada."))
            return redirect("events:invoice_detail", pk=event.pk, invoice_pk=invoice.pk)
    else:
        form = InvoiceItemForm(instance=item)
    return render(request, "events/invoice_item_form.html", {
        "form": form, "event": event, "invoice": invoice, "item": item,
    })


@login_required
def invoice_item_delete(request, pk, invoice_pk, item_pk):
    event = get_event_or_403(request.user, pk)
    if not request.user.can_manage_events:
        raise PermissionDenied
    invoice = get_object_or_404(Invoice, pk=invoice_pk, event=event)
    item = get_object_or_404(InvoiceItem, pk=item_pk, invoice=invoice)
    if invoice.is_locked:
        messages.error(request, _("Esta factura está anulada o pagada; no se puede modificar."))
        return redirect("events:invoice_detail", pk=event.pk, invoice_pk=invoice.pk)
    if request.method == "POST":
        item.delete()
        messages.success(request, _("Línea eliminada."))
    return redirect("events:invoice_detail", pk=event.pk, invoice_pk=invoice.pk)


@login_required
def invoice_print(request, pk, invoice_pk):
    event = get_event_or_403(request.user, pk)
    invoice = get_object_or_404(Invoice, pk=invoice_pk, event=event)
    return render(request, "events/invoice_print.html", {"event": event, "invoice": invoice})


@login_required
def report_itinerary(request, pk):
    event = get_event_or_403(request.user, pk)
    return render(request, "events/report_itinerary.html", {
        "event": event, "sessions": event.sessions.all(),
    })


@login_required
def report_tasks(request, pk):
    event = get_event_or_403(request.user, pk)
    tasks = list(event.tasks.select_related("assigned_to", "vendor", "supervisor", "chain").order_by(
        "assigned_to__first_name", "assigned_to__last_name", "category"
    ))
    responsibles = sorted({t.responsible_display for t in tasks})
    selected_responsible = request.GET.get("responsible", "")
    if selected_responsible:
        tasks = [t for t in tasks if t.responsible_display == selected_responsible]
    return render(request, "events/report_tasks.html", {
        "event": event, "tasks": tasks,
        "responsibles": responsibles, "selected_responsible": selected_responsible,
    })


@login_required
def report_status_history(request, pk):
    from apps.tasks.models import TaskStatusHistory

    event = get_event_or_403(request.user, pk)
    history = (
        TaskStatusHistory.objects.filter(task__event=event)
        .select_related("task", "changed_by")
        .order_by("-changed_at")
    )
    return render(request, "events/report_status_history.html", {"event": event, "history": history})


@login_required
def report_expenses(request, pk):
    event = get_event_or_403(request.user, pk)
    return render(request, "events/report_expenses.html", {"event": event, "expenses": event.expenses.all()})


@login_required
def report_vendors(request, pk):
    event = get_event_or_403(request.user, pk)
    bookings = event.event_vendors.select_related("vendor", "vendor__category")
    total_contract = bookings.aggregate(Sum("contract_amount"))["contract_amount__sum"] or 0
    total_paid = bookings.aggregate(Sum("deposit_paid"))["deposit_paid__sum"] or 0
    return render(request, "events/report_vendors.html", {
        "event": event, "bookings": bookings,
        "total_contract": total_contract,
        "total_paid": total_paid,
        "total_balance": total_contract - total_paid,
    })


@login_required
def report_vendors_excel(request, pk):
    from .xlsx_export import build_simple_workbook, workbook_response

    event = get_event_or_403(request.user, pk)
    bookings = event.event_vendors.select_related("vendor", "vendor__category")
    rows = [[
        b.vendor.name, b.vendor.category.name, b.get_status_display(),
        float(b.contract_amount), float(b.deposit_paid), float(b.balance_due),
    ] for b in bookings]
    totals = [
        str(_("TOTAL")), "", "",
        float(bookings.aggregate(Sum("contract_amount"))["contract_amount__sum"] or 0),
        float(bookings.aggregate(Sum("deposit_paid"))["deposit_paid__sum"] or 0),
        float(sum(b.balance_due for b in bookings)),
    ]
    wb = build_simple_workbook(
        f"{_('Proveedores')} - {event.name}",
        [str(_("Proveedor")), str(_("Categoría")), str(_("Estado")), str(_("Contrato")), str(_("Abonado")), str(_("Saldo"))],
        rows, totals, money_columns=[4, 5, 6],
    )
    return workbook_response(wb, f"proveedores_{event.pk}.xlsx")


@login_required
def report_meals(request, pk):
    event = get_event_or_403(request.user, pk)
    meals = event.meal_counts.all()
    groups = [
        (_("Equipo interno"), meals.filter(group=MealCount.GROUP_TEAM)),
        (_("Proveedores"), meals.filter(group=MealCount.GROUP_VENDOR)),
        (_("Imprevistos"), meals.filter(group=MealCount.GROUP_CONTINGENCY)),
    ]
    return render(request, "events/report_meals.html", {
        "event": event, "groups": groups,
        "grand_total": meals.aggregate(Sum("amount"))["amount__sum"] or 0,
    })


@login_required
def report_meals_excel(request, pk):
    from .xlsx_export import build_simple_workbook, workbook_response

    event = get_event_or_403(request.user, pk)
    meals = event.meal_counts.all()
    rows = [[
        m.get_group_display(), m.target_name, m.meal_label, m.count, float(m.amount), m.notes,
    ] for m in meals]
    totals = [str(_("TOTAL")), "", "", "", float(meals.aggregate(Sum("amount"))["amount__sum"] or 0), ""]
    wb = build_simple_workbook(
        f"{_('Comidas')} - {event.name}",
        [str(_("Grupo")), str(_("Proveedor / Persona")), str(_("Comida")), str(_("Cantidad")), str(_("Monto")), str(_("Notas"))],
        rows, totals, money_columns=[5],
    )
    return workbook_response(wb, f"comidas_{event.pk}.xlsx")


def _minute_by_minute_groups(event, only_pending, date_from="", date_to="", time_from="", time_to="", guion_only=False):
    """Groups every itinerary session by its (company-maintained) section,
    in section order — not just Ceremonia/Recepción, whatever sections the
    event actually uses.

    `guion_only=True` trims each session's task list down to tasks marked
    `is_guion`, the same way `only_pending` trims it down to non-completed
    tasks — the two filters compose independently.

    The date/time filter considers both the itinerary session's own date/time
    and each related task's own due_date/due_time (a task can be linked to a
    session but scheduled on a different day). A session shows up if it
    matches on its own or has at least one matching task; when the filter is
    active, only the tasks that themselves match are listed under it."""
    from django.db.models import Prefetch

    from apps.tasks.models import Task

    date_from_d = dt_date.fromisoformat(date_from) if date_from else None
    date_to_d = dt_date.fromisoformat(date_to) if date_to else None
    time_from_t = dt_time.fromisoformat(time_from) if time_from else None
    time_to_t = dt_time.fromisoformat(time_to) if time_to else None
    has_date_filter = bool(date_from_d or date_to_d or time_from_t or time_to_t)

    def matches_date_time(d, t):
        if date_from_d and (not d or d < date_from_d):
            return False
        if date_to_d and (not d or d > date_to_d):
            return False
        if time_from_t and (not t or t < time_from_t):
            return False
        if time_to_t and (not t or t > time_to_t):
            return False
        return True

    sessions = event.sessions.select_related("section").prefetch_related(
        Prefetch("tasks", queryset=Task.objects.select_related("chain"))
    ).order_by("section__order", "date", "start_time")

    groups_by_section = {}
    order = []
    for session in sessions:
        session_tasks = session.tasks.all()
        if guion_only:
            session_tasks = [t for t in session_tasks if t.is_guion]
        if only_pending:
            session_tasks = [t for t in session_tasks if t.status != "completada"]

        if has_date_filter:
            session_matches = matches_date_time(session.date, session.start_time)
            matching_tasks = [t for t in session_tasks if matches_date_time(t.due_date, t.due_time)]
            if not session_matches and not matching_tasks:
                continue
            session.related_tasks = sorted(
                matching_tasks, key=lambda t: (t.due_date or dt_date.max, t.due_time or dt_time.max)
            )
        else:
            session.related_tasks = sorted(
                session_tasks, key=lambda t: (t.due_date or dt_date.max, t.due_time or dt_time.max)
            )

        section = session.section
        if section.pk not in groups_by_section:
            groups_by_section[section.pk] = {"section": section, "sessions": []}
            order.append(section.pk)
        groups_by_section[section.pk]["sessions"].append(session)
    return [groups_by_section[pk] for pk in order]


def _minute_by_minute_filters(request):
    only_pending = request.GET.get("solo_pendientes") == "1"
    section_filters = request.GET.getlist("seccion")
    guion_only = request.GET.get("guion") == "1"
    lang = request.GET.get("lang") or getattr(request, "LANGUAGE_CODE", None) or translation.get_language()
    date_from = request.GET.get("fecha_desde") or ""
    date_to = request.GET.get("fecha_hasta") or ""
    time_from = request.GET.get("hora_desde") or ""
    time_to = request.GET.get("hora_hasta") or ""
    return only_pending, section_filters, guion_only, lang, date_from, date_to, time_from, time_to


@login_required
def report_minute_by_minute(request, pk):
    event = get_event_or_403(request.user, pk)
    only_pending, section_filters, guion_only, lang, date_from, date_to, time_from, time_to = _minute_by_minute_filters(request)
    groups = _minute_by_minute_groups(event, only_pending, date_from, date_to, time_from, time_to, guion_only)
    available_sections = [g["section"] for g in groups]

    section_links = []
    for section in available_sections:
        pk_str = str(section.pk)
        selected = pk_str in section_filters
        qd = request.GET.copy()
        current = qd.getlist("seccion")
        qd.setlist("seccion", [v for v in current if v != pk_str] if selected else current + [pk_str])
        section_links.append({"section": section, "selected": selected, "url": f"?{qd.urlencode()}"})
    qd_all = request.GET.copy()
    qd_all.pop("seccion", None)
    all_sections_url = f"?{qd_all.urlencode()}"

    if section_filters:
        groups = [g for g in groups if str(g["section"].pk) in section_filters]
    with translation.override(lang):
        return render(request, "events/report_minute_by_minute.html", {
            "event": event,
            "groups": groups,
            "section_links": section_links,
            "all_sections_url": all_sections_url,
            "only_pending": only_pending,
            "section_filters": section_filters,
            "guion_only": guion_only,
            "lang": lang,
            "date_from": date_from,
            "date_to": date_to,
            "time_from": time_from,
            "time_to": time_to,
        })


@login_required
def report_minute_by_minute_excel(request, pk):
    from .xlsx_export import build_simple_workbook, workbook_response
    from openpyxl.styles import Font

    event = get_event_or_403(request.user, pk)
    only_pending, section_filters, guion_only, lang, date_from, date_to, time_from, time_to = _minute_by_minute_filters(request)
    groups = _minute_by_minute_groups(event, only_pending, date_from, date_to, time_from, time_to, guion_only)
    if section_filters:
        groups = [g for g in groups if str(g["section"].pk) in section_filters]

    with translation.override(lang):
        rows = []
        itinerary_row_numbers = []
        for group in groups:
            section_name = group["section"].name
            for session in group["sessions"]:
                hora = session.start_time.strftime("%H:%M")
                if session.end_time:
                    hora += f" - {session.end_time.strftime('%H:%M')}"
                rows.append([
                    section_name, str(session.date), hora,
                    f"{_('Itinerario')}: {session.title}", session.venue_name or "", "", session.notes or "",
                ])
                itinerary_row_numbers.append(len(rows))
                for task in session.related_tasks:
                    task_hora = task.due_time.strftime("%H:%M") if task.due_time else ""
                    rows.append([
                        "", str(task.due_date) if task.due_date else "", task_hora,
                        f"{task.title} ({task.get_status_display()})", "", task.responsible_initials_display, "",
                    ])

        headers = [
            str(_("Sección")), str(_("Fecha")), str(_("Hora")),
            str(_("Itinerario")), str(_("Lugar")), str(_("Responsable")), str(_("Notas")),
        ]

        wb = build_simple_workbook(
            f"{_('Minuto a minuto')} - {event.name}",
            headers,
            rows,
        )
        ws = wb.active
        header_offset = ws.max_row - len(rows)
        for row_num in itinerary_row_numbers:
            for col in range(1, len(headers) + 1):
                ws.cell(row=header_offset + row_num, column=col).font = Font(bold=True)
    return workbook_response(wb, f"minuto_a_minuto_{event.pk}.xlsx")


def _wedding_party_report_list_types(request, event):
    """Returns (all_list_types_with_member_list, filtered_list_types, selected_pks)
    — 'lista' may be repeated in the querystring to report on several lists
    at once; with none selected, every list is included (unfiltered)."""
    create_default_wedding_party_list_types(event.company)
    all_list_types = list(WeddingPartyListType.objects.filter(company=event.company))
    members = event.wedding_party_members.select_related("list_type")
    for list_type in all_list_types:
        list_type.member_list = members.filter(list_type=list_type)
    selected = request.GET.getlist("lista")
    list_types = [lt for lt in all_list_types if str(lt.pk) in selected] if selected else all_list_types
    return all_list_types, list_types, selected


@login_required
def report_wedding_party(request, pk):
    event = get_event_or_403(request.user, pk)
    all_list_types, list_types, selected = _wedding_party_report_list_types(request, event)
    has_any_members = any(lt.member_list for lt in list_types)

    list_links = []
    for lt in all_list_types:
        pk_str = str(lt.pk)
        is_selected = pk_str in selected
        qd = request.GET.copy()
        current = qd.getlist("lista")
        qd.setlist("lista", [v for v in current if v != pk_str] if is_selected else current + [pk_str])
        list_links.append({"list_type": lt, "selected": is_selected, "url": f"?{qd.urlencode()}"})
    qd_all = request.GET.copy()
    qd_all.pop("lista", None)
    all_lists_url = f"?{qd_all.urlencode()}"

    return render(request, "events/report_wedding_party.html", {
        "event": event, "list_types": list_types, "has_any_members": has_any_members,
        "list_links": list_links, "all_lists_url": all_lists_url, "selected_lists": selected,
    })


@login_required
def report_wedding_party_excel(request, pk):
    from .xlsx_export import build_simple_workbook, workbook_response

    event = get_event_or_403(request.user, pk)
    _all_list_types, list_types, _selected = _wedding_party_report_list_types(request, event)

    rows = [
        [lt.name, m.order, m.name, m.role_description, m.quantity, m.table_number]
        for lt in list_types for m in lt.member_list
    ]
    wb = build_simple_workbook(
        f"{_('Cortejo nupcial')} - {event.name}",
        [
            str(_("Lista")), str(_("Orden")), str(_("Nombre")), str(_("Rol / descripción")),
            str(_("Cantidad")), str(_("Mesa")),
        ],
        rows,
    )
    return workbook_response(wb, f"cortejo_{event.pk}.xlsx")
