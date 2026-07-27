"""Parses the 'minuto a minuto' (run-of-show) spreadsheets planners already use
into plain dicts, so a view can turn them into EventSession rows.

Kept framework-free (no Django imports) so it can be unit-tested with just openpyxl.
"""
import re
from datetime import datetime, time as dt_time

import openpyxl


def _norm(value):
    return re.sub(r"\s+", " ", str(value or "")).strip().lower()


def _locate_header(ws, required_keywords, max_scan=10):
    best = (None, {})
    for row in ws.iter_rows(min_row=1, max_row=min(max_scan, ws.max_row)):
        cols = {}
        for cell in row:
            val = _norm(cell.value)
            if not val:
                continue
            for kw in required_keywords:
                if kw in val and kw not in cols:
                    cols[kw] = cell.column
        if len(cols) > len(best[1]):
            best = (row[0].row, cols)
    if best[0] is not None and len(best[1]) >= 2:
        return best
    return None, None


_AMPM_RE = re.compile(r"^\s*(\d{1,2}):(\d{2})\s*([AaPp][Mm])\s*$")
_HHMM_RE = re.compile(r"^\s*(\d{1,2}):(\d{2})\s*$")


def _parse_time_text(text):
    text = str(text).strip()
    m = _AMPM_RE.match(text)
    if m:
        hour, minute, ampm = int(m.group(1)), int(m.group(2)), m.group(3).lower()
        if ampm == "pm" and hour != 12:
            hour += 12
        if ampm == "am" and hour == 12:
            hour = 0
        if 0 <= hour <= 23 and 0 <= minute <= 59:
            return dt_time(hour, minute)
        return None
    m = _HHMM_RE.match(text)
    if m:
        hour, minute = int(m.group(1)), int(m.group(2))
        if 0 <= hour <= 23 and 0 <= minute <= 59:
            return dt_time(hour, minute)
    return None


def _cell_time(value):
    if isinstance(value, dt_time):
        return value
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, str) and value.strip():
        return _parse_time_text(value)
    return None


def _cell_str(ws, row, col):
    if col is None:
        return ""
    value = ws.cell(row=row, column=col).value
    if value is None:
        return ""
    return str(value).strip()


_GUEST_HEADER_KEYWORDS = [
    "nombre", "apellido", "mesa", "invita", "sexo", "rsvp",
    "familia", "plato", "restricciones", "anotaciones", "nota",
]
_GUEST_HEADER_REQUIRED = {"nombre", "apellido"}


_GUEST_HEADER_KEYWORDS_BY_LENGTH = sorted(_GUEST_HEADER_KEYWORDS, key=len, reverse=True)


def _locate_guest_header(ws, max_scan=10):
    """Same left-to-right, case-insensitive column scan as _locate_header, but only
    'nombre'/'apellido' are mandatory — the rest are picked up when present so
    older/simpler exports (missing some of the newer columns) still import.
    Each cell is matched against the *longest* keyword it contains (checked first),
    not just the first one found — this keeps 'nota' from wrongly claiming the
    'anotaciones' column (which contains 'nota' as a substring) regardless of
    which of the two columns comes first in the sheet."""
    best = (None, {})
    for row in ws.iter_rows(min_row=1, max_row=min(max_scan, ws.max_row)):
        cols = {}
        for cell in row:
            val = _norm(cell.value)
            if not val:
                continue
            for kw in _GUEST_HEADER_KEYWORDS_BY_LENGTH:
                if kw in val and kw not in cols:
                    cols[kw] = cell.column
                    break
        if len(cols) > len(best[1]):
            best = (row[0].row, cols)
    if best[0] is not None and _GUEST_HEADER_REQUIRED.issubset(best[1]):
        return best
    return None, None


_INVITA_MAP = {
    "novio": "novio", "groom": "novio",
    "novia": "novia", "bride": "novia",
    "ambos": "ambos", "both": "ambos", "los dos": "ambos",
}

_SEXO_MAP = {
    "hombre": "hombre", "h": "hombre", "masculino": "hombre", "male": "hombre", "man": "hombre", "m": "hombre",
    "mujer": "mujer", "mujeres": "mujer", "femenino": "mujer", "female": "mujer", "woman": "mujer", "f": "mujer",
    "nino": "nino", "niño": "nino", "boy": "nino", "chico": "nino",
    "nina": "nina", "niña": "nina", "girl": "nina", "chica": "nina",
}


def _map_invita(raw):
    text = _norm(raw)
    if not text:
        return ""
    if text in _INVITA_MAP:
        return _INVITA_MAP[text]
    has_novio = "novio" in text or "groom" in text
    has_novia = "novia" in text or "bride" in text
    if "ambos" in text or "both" in text or (has_novio and has_novia):
        return "ambos"
    if has_novio:
        return "novio"
    if has_novia:
        return "novia"
    return ""


def _map_sexo(raw):
    return _SEXO_MAP.get(_norm(raw), "")


_RELATIONSHIP_MAP = {
    "groom": "Novio", "bride": "Novia",
    "parent": "Padre/Madre",
    "close-relative": "Familiar cercano", "relative": "Familiar",
    "friend": "Amigo/a",
    "bridesmaid": "Dama", "best-man": "Padrino", "maid-of-honor": "Madrina",
}


def _map_relationship(raw):
    text = _norm(raw)
    if not text:
        return ""
    return _RELATIONSHIP_MAP.get(text, raw.strip())


class ParsedGuestRow:
    def __init__(self, first_name, last_name="", table_number="", invita="", sexo="",
                 family="", rsvp="", main_dish="", dietary_restrictions="", notes="",
                 relationship=""):
        self.first_name = first_name
        self.last_name = last_name
        self.table_number = table_number
        self.invita = invita
        self.sexo = sexo
        self.family = family
        self.rsvp = rsvp
        self.main_dish = main_dish
        self.dietary_restrictions = dietary_restrictions
        self.notes = notes
        self.relationship = relationship

    @property
    def name(self):
        return f"{self.first_name} {self.last_name}".strip()


def parse_guest_list(file_obj):
    """Format: columns 'nombre', 'apellido' (required), plus 'mesa', 'invita' (novio/
    novia/ambos), 'sexo' (hombre/mujer/niño/niña), 'rsvp', 'familia', 'plato principal',
    'restricciones', 'anotaciones' and 'nota' (role/relationship, e.g. groom/bride/
    relative/friend/bridesmaid) when present — as exported by wedding-planning tools
    (e.g. Wedding Assistant's guestlist export). Summary/footer rows (totals, 'powered
    by', etc.) have no 'nombre' and are skipped."""
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.worksheets[0]
    header_row, cols = _locate_guest_header(ws)
    if header_row is None:
        raise ValueError(
            "No se encontró una fila de encabezado con las columnas 'nombre' y 'apellido'."
        )

    name_col = cols.get("nombre")
    last_col = cols.get("apellido")
    mesa_col = cols.get("mesa")
    invita_col = cols.get("invita")
    sexo_col = cols.get("sexo")
    rsvp_col = cols.get("rsvp")
    familia_col = cols.get("familia")
    plato_col = cols.get("plato")
    restricciones_col = cols.get("restricciones")
    anotaciones_col = cols.get("anotaciones")
    nota_col = cols.get("nota")

    rows = []
    for row_num in range(header_row + 1, ws.max_row + 1):
        first_name = _cell_str(ws, row_num, name_col)
        if not first_name:
            continue

        rows.append(ParsedGuestRow(
            first_name=first_name,
            last_name=_cell_str(ws, row_num, last_col),
            table_number=_cell_str(ws, row_num, mesa_col),
            invita=_map_invita(_cell_str(ws, row_num, invita_col)),
            sexo=_map_sexo(_cell_str(ws, row_num, sexo_col)),
            family=_cell_str(ws, row_num, familia_col),
            rsvp=_cell_str(ws, row_num, rsvp_col),
            main_dish=_cell_str(ws, row_num, plato_col),
            dietary_restrictions=_cell_str(ws, row_num, restricciones_col),
            notes=_cell_str(ws, row_num, anotaciones_col),
            relationship=_map_relationship(_cell_str(ws, row_num, nota_col)),
        ))
    return rows


class ParsedSessionRow:
    def __init__(self, title, notes="", venue_name="", start_time=None, time_is_carried_over=False):
        self.title = title
        self.notes = notes
        self.venue_name = venue_name
        self.start_time = start_time
        self.time_is_carried_over = time_is_carried_over


TITLE_MAX_LEN = 150


def parse_minuto_a_minuto(file_obj):
    """Format: columns 'place', 'hour', 'details'. HOUR is often blank on rows that
    belong to the same moment as the row above (carried forward here); PLACE is
    used as given. Time text like '7:35PM' or a stray '?' is handled leniently."""
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.worksheets[0]
    header_row, cols = _locate_header(ws, ["place", "hour", "detail"])
    if header_row is None:
        raise ValueError(
            "No se encontró una fila de encabezado con las columnas 'place', 'hour' y 'details'."
        )

    place_col = cols.get("place")
    hour_col = cols.get("hour")
    details_col = cols.get("detail")

    rows = []
    last_time = None
    last_place = ""
    for row_num in range(header_row + 1, ws.max_row + 1):
        details = ws.cell(row=row_num, column=details_col).value if details_col else None
        details = str(details).strip() if details is not None else ""
        if not details:
            continue

        place_val = ws.cell(row=row_num, column=place_col).value if place_col else None
        place = str(place_val).strip() if place_val is not None else ""
        if place:
            last_place = place

        hour_val = ws.cell(row=row_num, column=hour_col).value if hour_col else None
        parsed_time = _cell_time(hour_val)
        carried_over = parsed_time is None
        if parsed_time is not None:
            last_time = parsed_time

        title = details if len(details) <= TITLE_MAX_LEN else details[: TITLE_MAX_LEN - 1] + "…"
        notes = details if len(details) > TITLE_MAX_LEN else ""

        rows.append(ParsedSessionRow(
            title=title,
            notes=notes,
            venue_name=last_place,
            start_time=last_time,
            time_is_carried_over=carried_over,
        ))
    return rows
