"""Parses the spreadsheets real wedding planners already use (guion completo,
task por persona) into plain dicts, so a view can turn them into Task rows.

Kept framework-free (no Django imports) so it can be unit-tested with just openpyxl.
"""
import re
from datetime import datetime, time as dt_time

import openpyxl


def normalize_text(value):
    return re.sub(r"\s+", " ", str(value or "")).strip().lower()


def _locate_header(ws, required_keywords, max_scan=10):
    """Scans the first rows for one containing most of the expected column keywords.
    Returns (header_row_number, {keyword: column_index}) or (None, None) if not found.
    """
    best = (None, {})
    for row in ws.iter_rows(min_row=1, max_row=min(max_scan, ws.max_row)):
        cols = {}
        for cell in row:
            val = normalize_text(cell.value)
            if not val:
                continue
            for kw in required_keywords:
                if kw in val and kw not in cols:
                    cols[kw] = cell.column
        if len(cols) > len(best[1]):
            best = (row[0].row, cols)
    if best[0] is not None and len(best[1]) >= max(2, len(required_keywords) - 2):
        return best
    return None, None


def _cell_str(ws, row, col):
    if col is None:
        return ""
    value = ws.cell(row=row, column=col).value
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _cell_time(ws, row, col):
    if col is None:
        return None
    value = ws.cell(row=row, column=col).value
    if isinstance(value, dt_time):
        return value
    if isinstance(value, datetime):
        return value.time()
    return None


def _cell_date(ws, row, col):
    if col is None:
        return None
    value = ws.cell(row=row, column=col).value
    if isinstance(value, datetime):
        return value.date()
    return None


def _is_truthy_done(value):
    if isinstance(value, bool):
        return value
    text = normalize_text(value)
    return text in ("true", "si", "sí", "x", "hecho", "1", "done", "completado", "completada")


class ParsedTaskRow:
    def __init__(self, title, responsible_name="", due_date=None, due_time=None, category="",
                 supplier_hint="", location=""):
        self.title = title
        self.responsible_name = responsible_name
        self.due_date = due_date
        self.due_time = due_time
        self.category = category
        # Free-text hints from formats that name a vendor *category* (e.g. "suplidor
        # flores") rather than a specific registered vendor — kept as context on the
        # task instead of used to auto-assign, since several vendors of the same
        # category can exist and the real one gets picked later.
        self.supplier_hint = supplier_hint
        self.location = location
        self.done = False


def parse_task_per_person(file_obj):
    """Format: columns 'responsible', 'task', 'hecho' (done checkbox)."""
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.worksheets[0]
    header_row, cols = _locate_header(ws, ["responsible", "task", "hecho"])
    if header_row is None:
        raise ValueError(
            "No se encontró una fila de encabezado con las columnas 'responsible', 'task' y 'hecho'."
        )

    rows = []
    for row_num in range(header_row + 1, ws.max_row + 1):
        title = _cell_str(ws, row_num, cols.get("task"))
        responsible = _cell_str(ws, row_num, cols.get("responsible"))
        if not title:
            continue
        parsed = ParsedTaskRow(title=title, responsible_name=responsible)
        done_col = cols.get("hecho")
        parsed.done = _is_truthy_done(ws.cell(row=row_num, column=done_col).value) if done_col else False
        rows.append(parsed)
    return rows


def parse_guion_completo(file_obj):
    """Format: columns 'fecha', 'hora', 'responsible', 'ubicacion', 'descripcion'.

    'fecha' in the source sheet is really the date the item was logged, not a hard
    deadline — it's used here as due_date since there's no better field, but rows
    commonly have no responsible/hora yet (still being planned).
    """
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.worksheets[0]
    header_row, cols = _locate_header(
        ws, ["fecha", "hora", "responsible", "ubicacion", "descripcion"]
    )
    if header_row is None:
        raise ValueError(
            "No se encontró una fila de encabezado con las columnas de fecha/hora/responsable/ubicación/descripción."
        )

    rows = []
    for row_num in range(header_row + 1, ws.max_row + 1):
        title = _cell_str(ws, row_num, cols.get("descripcion"))
        if not title:
            continue
        parsed = ParsedTaskRow(
            title=title,
            responsible_name=_cell_str(ws, row_num, cols.get("responsible")),
            due_date=_cell_date(ws, row_num, cols.get("fecha")),
            due_time=_cell_time(ws, row_num, cols.get("hora")),
            category=_cell_str(ws, row_num, cols.get("ubicacion")),
        )
        rows.append(parsed)
    return rows


def parse_guion_final(file_obj):
    """Format: columns 'fecha', 'hora', 'responsible', 'actividad', 'proveedor',
    'ubicacion', 'descripcion'. 'responsible' may list several people separated by
    '/' (e.g. 'Novia / Planner') — each becomes its own row so every person can
    check off their own copy of the task independently.
    """
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.worksheets[0]
    header_row, cols = _locate_header(
        ws, ["fecha", "hora", "responsible", "actividad", "proveedor", "ubicacion", "descripcion"]
    )
    if header_row is None:
        raise ValueError(
            "No se encontró una fila de encabezado con las columnas de fecha/hora/responsable/"
            "actividad/proveedor/ubicación/descripción."
        )

    rows = []
    for row_num in range(header_row + 1, ws.max_row + 1):
        title = _cell_str(ws, row_num, cols.get("descripcion"))
        if not title:
            continue

        due_date = _cell_date(ws, row_num, cols.get("fecha"))
        due_time = _cell_time(ws, row_num, cols.get("hora"))
        category = _cell_str(ws, row_num, cols.get("actividad"))
        supplier_hint = _cell_str(ws, row_num, cols.get("proveedor"))
        location = _cell_str(ws, row_num, cols.get("ubicacion"))

        responsible_raw = _cell_str(ws, row_num, cols.get("responsible"))
        names = [n.strip() for n in responsible_raw.split("/") if n.strip()] or [""]
        for name in names:
            rows.append(ParsedTaskRow(
                title=title, responsible_name=name, due_date=due_date, due_time=due_time,
                category=category, supplier_hint=supplier_hint, location=location,
            ))
    return rows


IGNORED_ITINERARY_LOCATIONS = {"escritorio", ""}
IGNORED_ITINERARY_CATEGORIES = {"general", "geneal", "personal", ""}

# Mirrors apps.events.models.EventSession.SECTION_* values (kept as plain strings
# here since this module stays framework-free / Django-import-free).
SECTION_CEREMONIA = "ceremonia"
SECTION_RECEPCION = "recepcion"
SECTION_MONTAJE = "montaje"
SECTION_DESMONTAJE = "desmontaje"
SECTION_OTRO = "otro"

_CEREMONIA_KEYWORDS = ["iglesia", "church", "ceremonia"]
_RECEPCION_KEYWORDS = [
    "recepcion", "reception", "lobby", "cocktail", "coctel", "cocktal",
    "salon", "banquete", "brindis", "baile", "cena", "dinner",
]
_MONTAJE_KEYWORDS = ["montaje", "set up", "setup"]
_DESMONTAJE_KEYWORDS = ["desmontaje", "recogida", "load out", "loadout"]


def classify_section(location, category):
    """Best-effort keyword match on location/category text (e.g. 'Iglesia',
    'Iglesia/Venue', 'Iglesia*' -> Ceremonia; 'Recepción', 'Lobby', 'Cocktail' ->
    Recepción) so the itinerary proposal from guion_final groups activities the
    way a planner already thinks about them."""
    text = normalize_text(f"{location} {category}")
    if any(kw in text for kw in _CEREMONIA_KEYWORDS):
        return SECTION_CEREMONIA
    # Checked before montaje: "desmontaje" contains "montaje" as a substring.
    if any(kw in text for kw in _DESMONTAJE_KEYWORDS):
        return SECTION_DESMONTAJE
    if any(kw in text for kw in _MONTAJE_KEYWORDS):
        return SECTION_MONTAJE
    if any(kw in text for kw in _RECEPCION_KEYWORDS):
        return SECTION_RECEPCION
    return SECTION_OTRO


def propose_itinerary_from_rows(rows, center_date, day_window=3):
    """Groups guion_final-style rows (with due_date/due_time/category/location) into
    itinerary-activity proposals restricted to the days around the event (rehearsal,
    montaje, the event day itself, desmontaje). This is only ever a *proposal* — a
    human reviews and edits it before any EventSession gets created from it, since
    the source HORA values are sometimes just sequential placeholders, not real times.
    """
    if not center_date:
        return []
    groups = {}
    for row in rows:
        if not row.due_date or abs((row.due_date - center_date).days) > day_window:
            continue
        if normalize_text(row.location) in IGNORED_ITINERARY_LOCATIONS:
            continue
        if normalize_text(row.category) in IGNORED_ITINERARY_CATEGORIES:
            continue
        key = (row.due_date, row.location.strip(), row.category.strip())
        group = groups.setdefault(key, {
            "due_date": row.due_date, "location": row.location.strip(),
            "category": row.category.strip(), "times": [], "sample_titles": [],
        })
        if row.due_time:
            group["times"].append(row.due_time)
        if row.title not in group["sample_titles"] and len(group["sample_titles"]) < 3:
            group["sample_titles"].append(row.title)

    proposals = []
    for group in groups.values():
        if not group["times"]:
            continue
        proposals.append({
            "due_date": group["due_date"],
            "start_time": min(group["times"]),
            "venue_name": group["location"],
            "title": group["category"],
            "notes": "; ".join(group["sample_titles"]),
            "section": classify_section(group["location"], group["category"]),
        })
    section_order = {
        SECTION_CEREMONIA: 0, SECTION_RECEPCION: 1, SECTION_MONTAJE: 2,
        SECTION_DESMONTAJE: 3, SECTION_OTRO: 4,
    }
    proposals.sort(key=lambda p: (section_order[p["section"]], p["due_date"], p["start_time"]))
    return proposals


def match_user_by_name(name, candidate_users):
    """Best-effort match of a free-text name (e.g. 'FULCAR', 'Paige Smith') against
    a list of Users, by comparing to their full name / username. Returns None if
    nothing looks like a reasonable match (kept deliberately conservative: partial
    matches only count when the raw name has at least one whole word in common)."""
    name_norm = normalize_text(name)
    if not name_norm:
        return None
    name_words = set(name_norm.split())
    for user in candidate_users:
        full_name = normalize_text(user.get_full_name())
        username = normalize_text(user.username)
        if name_norm == full_name or name_norm == username:
            return user
        if full_name and name_words & set(full_name.split()):
            return user
    return None


def match_vendor_by_name(name, candidate_vendors):
    """Same best-effort matching as match_user_by_name, but against registered
    Vendor records (e.g. 'FULCAR' matching a transportation vendor already on file)."""
    name_norm = normalize_text(name)
    if not name_norm:
        return None
    name_words = set(name_norm.split())
    for vendor in candidate_vendors:
        vendor_name = normalize_text(vendor.name)
        if name_norm == vendor_name:
            return vendor
        if vendor_name and name_words & set(vendor_name.split()):
            return vendor
    return None
